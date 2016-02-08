from pybrain.tools.xml.networkreader import NetworkReader
from pybrain.tools.shortcuts import buildNetwork
from pybrain.datasets import SupervisedDataSet
from pybrain.supervised.trainers import RPropMinusTrainer
from pybrain.tools.xml.networkwriter import NetworkWriter
from openpyxl import load_workbook


def createAndTrainNetworkFromFile(curs_filename, count_input_samples, count_samples, net_filename, count_layers=33,
                          count_outputs=1, max_epochs=15000, min_epochs=300):
    net = buildNetwork(count_input_samples, count_layers, count_outputs)
    ds = SupervisedDataSet(count_input_samples, count_outputs)
    wb = load_workbook(filename=curs_filename)
    ws = wb.active
    for i in range(0, count_samples):
        loaded_data = []
        for j in range(0, count_input_samples + 1):
            loaded_data.append(round(float(ws.cell(row=i+1, column=j+1).value), 4))
            #ds.addSample(loaded_data[:-1], loaded_data[-1])
        #print loaded_data[:-1], loaded_data[-1]
        ds.addSample(loaded_data[:-1], loaded_data[-1])
    trainer = RPropMinusTrainer(net, verbose=True)
    trainer.setData(ds)
    a = trainer.trainUntilConvergence(maxEpochs=max_epochs, continueEpochs=min_epochs, validationProportion=0.15)
    net_filename = net_filename[:-4]+str(a[0][-1])+'.xml'
    NetworkWriter.writeToFile(net, net_filename)
    result_list = [a, net_filename]
    return result_list


def createAndTrainNetworkFromList(train_list, count_input_samples, net_filename, count_layers=33,
                          count_outputs=1, max_epochs=15000, min_epochs=300):
    net = buildNetwork(count_input_samples, count_layers, count_outputs)
    ds = SupervisedDataSet(count_input_samples, count_outputs)
    count_samples = len(train_list)
    for i in range(0, count_samples):
        ds.addSample(train_list[i][:-count_outputs], train_list[i][-count_outputs])
    trainer = RPropMinusTrainer(net, verbose=True)
    trainer.setData(ds)
    a = trainer.trainUntilConvergence(maxEpochs=max_epochs, continueEpochs=min_epochs, validationProportion=0.15)
    net_filename = net_filename[:-4]+str(a[0][-1])+'.xml'
    NetworkWriter.writeToFile(net, net_filename)
    result_list = [a, net_filename]
    return result_list


def loadNet(net_filename):
    net = NetworkReader.readFrom(net_filename)
    return net


def trainNetwork(net, sample_list, validate_list, net_filename, max_epochs=5500, min_epochs=300):
    count_input_samples = len(sample_list)
    count_outputs = len(validate_list)
    ds = SupervisedDataSet(count_input_samples, count_outputs)
    ds.addSample(sample_list, validate_list)
    trainer = RPropMinusTrainer(net, verbose=True)
    trainer.setData(ds)
    trainer.trainUntilConvergence(maxEpochs=max_epochs, continueEpochs=min_epochs)
    NetworkWriter.writeToFile(net, net_filename)
    return net