# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import csv
from datetime import datetime, date, timedelta
from ..models import Quote
from getQuotesData import getUSDQuotesOnDate


def migrateDataFromXlsx(xlsx_filename, count_inputs=6):
    wb = load_workbook(xlsx_filename)
    ws = wb.active
    count_samples = len(ws.rows)
    input_data = []
    output_data = []
    samples_date = []
    for i in range(1, count_samples+1):
        inputs_data = ""
        for j in range(1, count_inputs+1):
            inputs_data += str(ws.cell(row=i, column=j).value) + ", "
        input_data.append(inputs_data[:-2])
        output_data.append(ws.cell(row=i, column=count_inputs+1).value)
        samples_date.append(ws.cell(row=i, column=count_inputs+2).value)
    return {'input_data': input_data, 'output_data': output_data, 'samples_date': samples_date}


def migrateOilQuotesFromFile(filename):
    f = open(filename)
    reader = csv.reader(f)
    curs_date = '01.01.2015'
    count_migrate_qoutes = 0
    quote_type = 'brent'
    for row_index, row in enumerate(reader):
        if curs_date != row[0].split(';')[0]:
            if row[0].split(';')[1] == '17:00:00':
                quote_date = datetime.strptime(row[0].split(';')[0], "%d.%m.%Y")
                try:
                    q = Quote.objects.get(quote_date=quote_date, quote_type=quote_type)
                except Quote.DoesNotExist:
                    print "Quote on date" + str(quote_date) + "doesn't exist"
                    if date.weekday(quote_date) == 4:
                        q_sat = Quote(quote_value=round(float(row[0].split(';')[5]), 4), quote_type=quote_type,
                                      quote_date=quote_date+timedelta(days=1))
                        q_sat.save()
                        q_sun = Quote(quote_value=round(float(row[0].split(';')[5]), 4), quote_type=quote_type,
                                      quote_date=quote_date+timedelta(days=2))
                    q_sun.save()
                    q = Quote(quote_value=round(float(row[0].split(';')[5]), 4), quote_type=quote_type, quote_date=quote_date)
                    q.save()
                    count_migrate_qoutes += 1
    return count_migrate_qoutes


def migrateUSDQuotesFromCB(start_date):
    date_delta = date.today() - datetime.date(datetime.strptime(start_date, '%d.%m.%Y'))
    quote_type = 'usd'
    count_migrate_qoutes = 0
    for i in range(0, date_delta.days):
        quote_date = (date.today() - timedelta(days=i)).__str__()
        try:
            q = Quote.objects.get(quote_date=quote_date, quote_type=quote_type)
        except Quote.DoesNotExist:
            quote_value = getUSDQuotesOnDate(quote_date)
            q = Quote(quote_value=round(float(quote_value), 4), quote_type=quote_type, quote_date=quote_date)
            q.save()
    return count_migrate_qoutes

