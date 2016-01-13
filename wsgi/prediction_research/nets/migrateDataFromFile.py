# -*- coding: utf-8 -*-
from openpyxl import load_workbook


def migrateDataFromXlsx(xlsx_filename, count_inputs=6):
    wb = load_workbook(xlsx_filename)
    ws = wb.active
    count_samples = len(ws.rows)
    input_data = []
    output_data = []
    samples_date = []
    for i in range(1, count_samples+1):
        inputs_data = ""
        for j in range(1, count_inputs+1):
            inputs_data += str(ws.cell(row=i, column=j).value) + ", "
        input_data.append(inputs_data[:-2])
        output_data.append(ws.cell(row=i, column=count_inputs+1).value)
        samples_date.append(ws.cell(row=i, column=count_inputs+2).value)
    return {'input_data': input_data, 'output_data': output_data, 'samples_date': samples_date}
