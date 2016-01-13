from openpyxl import load_workbook, Workbook
from datetime import date, timedelta, datetime
from getQuotesData import getUSDQuotesOnDate, getCurrentOilQuotes
import csv


def makeTrainFile(train_filename, count_samples, sample_length):
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Data can be assigned directly to cells
    curslist = []

    for i in range(0, count_samples):
        if len(curslist) > 0:
            curslist.pop()   # remove date
            curslist.pop(4)  # remove last day value
            start = sample_length - 1
        else:
            start = 0
        curslist.insert(5, (date.today() - timedelta(days=i+1)).__str__())
        for j in range(start, sample_length):
            day = (date.today() - timedelta(days=j+i)).__str__()
            c = getUSDQuotesOnDate(day)
            curslist.insert(0, c)
        print curslist
        ws.append(curslist)
    # Python types will automatically be converted
    # Save the file
    wb.save(train_filename)


def mergeDollarAndOilCurs(dollar_file, oil_file, output_file, count_samples, start_column_index, count_shifts):
    wb_dollar_curs = load_workbook(filename = dollar_file)
    ws_dollar_curs = wb_dollar_curs.active
    f = open(oil_file)
    reader = csv.reader(f)
    reversed_csv_value_list = []
    reversed_date_list = []
    curs_date = '01.01.2015'
    for row_index, row in enumerate(reader):
        if curs_date != row[0].split(';')[0]:
            if row[0].split(';')[1] == '17:00:00':
                reversed_csv_value_list.append(round(float(row[0].split(';')[5]), 4))
                reversed_date_list.append(datetime.strptime(row[0].split(';')[0], "%d.%m.%Y"))
    csv_value_list = reversed_csv_value_list[::-1]
    date_list = reversed_date_list[::-1]
    print csv_value_list
    print date_list
    k = 0
    for i in range(1, count_samples + 1):
        w = 0
        oil_date = date_list[k]
        dollar_date = datetime.strptime(ws_dollar_curs.cell(row=i, column=start_column_index+1).value, "%Y-%m-%d")
        print oil_date
        print dollar_date
        print date.weekday(dollar_date)
        if oil_date == dollar_date:
            for j in range(count_shifts-1, -1, -1):
                print j
                ws_dollar_curs.cell(row=i, column=start_column_index+count_shifts+j).value = ws_dollar_curs.cell(row=i, column=start_column_index+j).value
                ws_dollar_curs.cell(row=i, column=start_column_index+j).value = csv_value_list[k+1-j-w]
            k += 1
        if date.weekday(dollar_date) >= 5:
            if date.weekday(dollar_date) == 5:
                w = 1
            if date.weekday(dollar_date) == 6:
                w = 2
            k += w
            for j in range(0, count_shifts):
                ws_dollar_curs.cell(row=i, column=start_column_index+count_shifts+j).value = ws_dollar_curs.cell(row=i, column=start_column_index+j).value
                ws_dollar_curs.cell(row=i, column=start_column_index+j).value = csv_value_list[k-w]
            k -= w
    wb_dollar_curs.save(output_file)


def updateTrainFile(count_dollar_samples, count_oil_samples, train_file):
    wb = load_workbook(filename=train_file)
    ws = wb.active
    count_rows = len(ws.rows)
    ws.append(ws.rows[-1])
    for i in range(count_rows - 1, 0, -1):
        print i
        print ws.rows[i]
        for j in range(1, len(ws.rows[i])):
            ws.cell(row=i+1,column=j).value = ws.cell(row=i,column=j).value
    dollar_today_value = getUSDQuotesOnDate(date.today().__str__())
    oil_today_value = getCurrentOilQuotes()
    wb.save('simple.xlsx')
    return 1

